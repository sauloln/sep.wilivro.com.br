import os
import csv
import io
from datetime import datetime, timedelta
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, session, Response, jsonify
from sqlalchemy import func, desc
from sqlalchemy.orm import aliased

from config import Config
from models import (
    db, Configuracao,
    Estado, Municipio, Local,
    Cliente, Projeto,
    PatrimonioItem, PatrimonioMov,
    Deposito, Produto, EstoqueSaldo, EstoqueMov
)

DEFAULT_UNIDADES = ["UN", "CX", "PCT", "KG", "G", "L", "ML", "M", "M2", "M3"]


def create_app():
    app = Flask(__name__, instance_relative_config=True)
    app.config.from_object(Config)

    os.makedirs(app.instance_path, exist_ok=True)

    # SQLite em instance/
    if app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite:///"):
        app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(app.instance_path, "sep.db")

    upload_folder = app.config["UPLOAD_FOLDER"]
    if not os.path.isabs(upload_folder):
        upload_folder = os.path.join(app.root_path, upload_folder)
    app.config["UPLOAD_FOLDER"] = upload_folder
    os.makedirs(upload_folder, exist_ok=True)

    db.init_app(app)

    # ---------------- Config helpers ----------------
    def get_config(chave, default=None):
        c = Configuracao.query.filter_by(chave=chave).first()
        return c.valor if c else default

    def set_config(chave, valor):
        c = Configuracao.query.filter_by(chave=chave).first()
        if not c:
            c = Configuracao(chave=chave, valor=str(valor))
            db.session.add(c)
        else:
            c.valor = str(valor)
        db.session.commit()

    @app.context_processor
    def inject_globals():
        return {"app_title": app.config.get("APP_TITLE", "SEP")}

    def admin_required(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not session.get("admin_ok"):
                return redirect(url_for("admin_login"))
            return f(*args, **kwargs)
        return wrapper

    # ---------------- Banco: migração leve (SQLite) ----------------
    def _ensure_sqlite_columns(_app):
        uri = _app.config.get("SQLALCHEMY_DATABASE_URI", "")
        if not uri.startswith("sqlite:///"):
            return

        from sqlalchemy import text

        def has_col(table, col):
            rows = db.session.execute(text(f"PRAGMA table_info({table})")).fetchall()
            return any(r[1] == col for r in rows)

        # patrimonio_item
        if not has_col("patrimonio_item", "projeto_atual_id"):
            db.session.execute(text("ALTER TABLE patrimonio_item ADD COLUMN projeto_atual_id INTEGER"))

        # patrimonio_mov
        if not has_col("patrimonio_mov", "projeto_id"):
            db.session.execute(text("ALTER TABLE patrimonio_mov ADD COLUMN projeto_id INTEGER"))

        if not has_col("patrimonio_mov", "qtd"):
            db.session.execute(text("ALTER TABLE patrimonio_mov ADD COLUMN qtd INTEGER NOT NULL DEFAULT 1"))

        db.session.commit()

    def next_patrimonio_codigo():
        last = PatrimonioItem.query.order_by(PatrimonioItem.id.desc()).first()
        nxt = (last.id + 1) if last else 1
        return f"PAT-{nxt:06d}"

    def saldo_get_or_create(produto_id, deposito_id):
        s = EstoqueSaldo.query.filter_by(produto_id=produto_id, deposito_id=deposito_id).first()
        if not s:
            s = EstoqueSaldo(produto_id=produto_id, deposito_id=deposito_id, saldo=0)
            db.session.add(s)
            db.session.flush()
        return s

    def apply_estoque_mov(tipo, produto_id, qtd, dep_origem_id=None, dep_destino_id=None, custo_unit=None, doc=None, obs=None):
        # Regras:
        # entrada: + destino
        # saida: - origem
        # transferencia: - origem + destino
        # ajuste: set saldo do destino para qtd (qtd = novo saldo)
        mv = EstoqueMov(
            tipo=tipo,
            produto_id=produto_id,
            qtd=qtd,
            deposito_origem_id=dep_origem_id,
            deposito_destino_id=dep_destino_id,
            custo_unit=custo_unit,
            documento_ref=doc,
            observacao=obs,
        )
        db.session.add(mv)
        db.session.flush()

        if tipo == "entrada":
            if not dep_destino_id:
                raise ValueError("Entrada exige depósito destino")
            s = saldo_get_or_create(produto_id, dep_destino_id)
            s.saldo += float(qtd)

        elif tipo == "saida":
            if not dep_origem_id:
                raise ValueError("Saída exige depósito origem")
            s = saldo_get_or_create(produto_id, dep_origem_id)
            s.saldo -= float(qtd)

        elif tipo == "transferencia":
            if not dep_origem_id or not dep_destino_id:
                raise ValueError("Transferência exige origem e destino")
            s1 = saldo_get_or_create(produto_id, dep_origem_id)
            s2 = saldo_get_or_create(produto_id, dep_destino_id)
            s1.saldo -= float(qtd)
            s2.saldo += float(qtd)

        elif tipo == "ajuste":
            if not dep_destino_id:
                raise ValueError("Ajuste exige depósito destino")
            s = saldo_get_or_create(produto_id, dep_destino_id)
            s.saldo = float(qtd)

        else:
            raise ValueError("Tipo inválido")

        db.session.commit()
        return mv

    # ---------------- Public/Login gate ----------------
    @app.before_request
    def _require_login_for_app():
        path = request.path
        if path.startswith("/static/"):
            return None
        if path in ("/admin/login", "/health", "/"):
            return None
        if path.startswith("/api/"):
            if not session.get("admin_ok"):
                return jsonify({"results": []}), 401
            return None
        if not session.get("admin_ok"):
            return redirect(url_for("admin_login"))
        return None

    @app.route("/")
    def index():
        if session.get("admin_ok"):
            return redirect(url_for("admin_dashboard"))
        return render_template("index.html")

    @app.route("/health")
    def health():
        return jsonify({"ok": True, "time": datetime.utcnow().isoformat() + "Z"})

    # ---------------- Admin Auth ----------------
    @app.route("/admin/login", methods=["GET", "POST"])
    def admin_login():
        if request.method == "POST":
            if request.form.get("username") == app.config["ADMIN_USERNAME"] and request.form.get("password") == app.config["ADMIN_PASSWORD"]:
                session["admin_ok"] = True
                flash("Login realizado.", "success")
                return redirect(url_for("admin_dashboard"))
            flash("Credenciais inválidas.", "danger")
        return render_template("admin_login.html")

    @app.route("/admin/logout")
    def admin_logout():
        session.pop("admin_ok", None)
        flash("Sessão encerrada.", "info")
        return redirect(url_for("index"))

    @app.route("/admin")
    @admin_required
    def admin_dashboard():
        mov_30d = EstoqueMov.query.filter(EstoqueMov.data >= datetime.utcnow() - timedelta(days=30)).count()
        kpis = {
            "patrimonio_total": PatrimonioItem.query.count(),
            "produtos_total": Produto.query.count(),
            "depositos_total": Deposito.query.count(),
            "mov_30d": mov_30d,
        }
        return render_template("admin_dashboard.html", kpis=kpis)

    # -------- APIs (autocomplete) --------
    @app.route("/api/municipios")
    @admin_required
    def api_municipios():
        q = (request.args.get("q") or "").strip()
        query = (db.session.query(Municipio.id, Municipio.nome, Estado.sigla)
                 .join(Estado, Municipio.estado_id == Estado.id))
        if q:
            like = f"%{q}%"
            parts = q.replace("-", " ").split()
            if len(parts) >= 2 and len(parts[-1]) == 2:
                uf = parts[-1].upper()
                name = " ".join(parts[:-1])
                query = query.filter(Estado.sigla == uf, Municipio.nome.ilike(f"%{name}%"))
            else:
                query = query.filter((Municipio.nome.ilike(like)) | (Estado.sigla.ilike(like)))
        rows = query.order_by(Estado.sigla, Municipio.nome).limit(30).all()
        return jsonify({"results": [{"id": r.id, "text": f"{r.nome} - {r.sigla} (IBGE {r.id})"} for r in rows]})

    @app.route("/api/locais")
    @admin_required
    def api_locais():
        q = (request.args.get("q") or "").strip()
        query = (db.session.query(Local.id, Local.nome, Municipio.nome.label("mun"), Estado.sigla.label("uf"))
                 .join(Municipio, Local.municipio_id == Municipio.id)
                 .join(Estado, Municipio.estado_id == Estado.id))
        if q:
            like = f"%{q}%"
            query = query.filter((Local.nome.ilike(like)) | (Municipio.nome.ilike(like)) | (Estado.sigla.ilike(like)))
        rows = query.order_by(Estado.sigla, Municipio.nome, Local.nome).limit(30).all()
        return jsonify({"results": [{"id": r.id, "text": f"{r.nome} — {r.mun}/{r.uf} (ID {r.id})"} for r in rows]})

    @app.route("/api/produtos")
    @admin_required
    def api_produtos():
        q = (request.args.get("q") or "").strip()
        query = db.session.query(Produto.id, Produto.sku, Produto.nome, Produto.tipo, Produto.unidade)
        if q:
            like = f"%{q}%"
            query = query.filter((Produto.nome.ilike(like)) | (Produto.sku.ilike(like)))
        rows = query.order_by(Produto.nome).limit(30).all()
        return jsonify({"results": [{"id": r.id, "text": f"{r.sku} — {r.nome} ({r.tipo}, {r.unidade})"} for r in rows]})

    @app.route("/api/depositos")
    @admin_required
    def api_depositos():
        q = (request.args.get("q") or "").strip()
        query = db.session.query(Deposito.id, Deposito.nome, Deposito.tipo)
        if q:
            like = f"%{q}%"
            query = query.filter((Deposito.nome.ilike(like)) | (Deposito.tipo.ilike(like)))
        rows = query.order_by(Deposito.nome).limit(30).all()
        return jsonify({"results": [{"id": r.id, "text": f"{r.nome} ({r.tipo}) — ID {r.id}"} for r in rows]})

    @app.route("/api/clientes")
    @admin_required
    def api_clientes():
        q = (request.args.get("q") or "").strip()
        query = db.session.query(Cliente.id, Cliente.nome)
        if q:
            like = f"%{q}%"
            query = query.filter(Cliente.nome.ilike(like))
        rows = query.order_by(Cliente.nome).limit(30).all()
        return jsonify({"results": [{"id": r.id, "text": f"{r.nome} (ID {r.id})"} for r in rows]})

    @app.route("/api/projetos")
    @admin_required
    def api_projetos():
        q = (request.args.get("q") or "").strip()
        query = db.session.query(Projeto.id, Projeto.nome)
        if q:
            like = f"%{q}%"
            query = query.filter(Projeto.nome.ilike(like))
        rows = query.order_by(Projeto.nome).limit(30).all()
        return jsonify({"results": [{"id": r.id, "text": f"{r.nome} (ID {r.id})"} for r in rows]})

    @app.route("/api/patrimonio_itens")
    @admin_required
    def api_patrimonio_itens():
        q = (request.args.get("q") or "").strip()
        query = db.session.query(PatrimonioItem.id, PatrimonioItem.codigo, PatrimonioItem.descricao)
        if q:
            like = f"%{q}%"
            query = query.filter((PatrimonioItem.codigo.ilike(like)) | (PatrimonioItem.descricao.ilike(like)))
        rows = query.order_by(PatrimonioItem.id.desc()).limit(30).all()
        return jsonify({"results": [{"id": r.id, "text": f"{r.codigo} — {r.descricao} (ID {r.id})"} for r in rows]})

    # -------- Locais --------
    @app.route("/admin/locais", methods=["GET", "POST"])
    @admin_required
    def admin_locais():
        if request.method == "POST":
            nome = (request.form.get("nome") or "").strip()
            tipo = (request.form.get("tipo") or "escritorio").strip()
            municipio_id = request.form.get("municipio_id", type=int)
            endereco = (request.form.get("endereco") or "").strip() or None

            if not nome or not municipio_id:
                flash("Informe nome e município (ID).", "warning")
                return redirect(url_for("admin_locais"))

            mun = Municipio.query.get(municipio_id)
            if not mun:
                flash("Município não encontrado. Rode o seed IBGE.", "danger")
                return redirect(url_for("admin_locais"))

            db.session.add(Local(nome=nome, tipo=tipo, municipio_id=municipio_id, endereco=endereco))
            db.session.commit()
            flash("Local criado.", "success")
            return redirect(url_for("admin_locais"))

        locais = (Local.query
                  .join(Municipio, Local.municipio_id == Municipio.id)
                  .join(Estado, Municipio.estado_id == Estado.id)
                  .order_by(Estado.sigla, Municipio.nome, Local.nome)
                  .all())
        return render_template("admin_locais.html", locais=locais)

    @app.route("/admin/locais/<int:local_id>/edit", methods=["GET", "POST"])
    @admin_required
    def admin_local_edit(local_id):
        local = Local.query.get_or_404(local_id)
        if request.method == "POST":
            local.nome = (request.form.get("nome") or "").strip()
            local.tipo = (request.form.get("tipo") or "escritorio").strip()
            mid = request.form.get("municipio_id", type=int)
            if not Municipio.query.get(mid):
                flash("Município inválido.", "danger")
                return redirect(url_for("admin_local_edit", local_id=local.id))
            local.municipio_id = mid
            local.endereco = (request.form.get("endereco") or "").strip() or None
            local.observacao = (request.form.get("observacao") or "").strip() or None
            db.session.commit()
            flash("Local atualizado.", "success")
            return redirect(url_for("admin_locais"))
        return render_template("admin_local_edit.html", local=local)

    @app.route("/admin/locais/<int:local_id>/delete", methods=["POST"])
    @admin_required
    def admin_local_delete(local_id):
        local = Local.query.get_or_404(local_id)
        db.session.delete(local)
        db.session.commit()
        flash("Local removido.", "info")
        return redirect(url_for("admin_locais"))

    # -------- Depósitos --------
    @app.route("/admin/depositos", methods=["GET", "POST"])
    @admin_required
    def admin_depositos():
        if request.method == "POST":
            nome = (request.form.get("nome") or "").strip()
            tipo = (request.form.get("tipo") or "interno").strip()
            local_id = request.form.get("local_id", type=int)
            local_id = local_id if local_id else None
            db.session.add(Deposito(nome=nome, tipo=tipo, local_id=local_id))
            db.session.commit()
            flash("Depósito criado.", "success")
            return redirect(url_for("admin_depositos"))
        depositos = Deposito.query.order_by(Deposito.nome).all()
        return render_template("admin_depositos.html", depositos=depositos)

    @app.route("/admin/depositos/<int:deposito_id>/edit", methods=["GET", "POST"])
    @admin_required
    def admin_deposito_edit(deposito_id):
        deposito = Deposito.query.get_or_404(deposito_id)
        if request.method == "POST":
            deposito.nome = (request.form.get("nome") or "").strip()
            deposito.tipo = (request.form.get("tipo") or "interno").strip()
            lid = request.form.get("local_id", type=int)
            deposito.local_id = lid if lid else None
            deposito.observacao = (request.form.get("observacao") or "").strip() or None
            db.session.commit()
            flash("Depósito atualizado.", "success")
            return redirect(url_for("admin_depositos"))
        return render_template("admin_deposito_edit.html", deposito=deposito)

    @app.route("/admin/depositos/<int:deposito_id>/delete", methods=["POST"])
    @admin_required
    def admin_deposito_delete(deposito_id):
        deposito = Deposito.query.get_or_404(deposito_id)
        db.session.delete(deposito)
        db.session.commit()
        flash("Depósito removido.", "info")
        return redirect(url_for("admin_depositos"))

    # -------- Produtos --------
    @app.route("/admin/produtos", methods=["GET", "POST"])
    @admin_required
    def admin_produtos():
        if request.method == "POST":
            sku = (request.form.get("sku") or "").strip()
            nome = (request.form.get("nome") or "").strip()
            tipo = (request.form.get("tipo") or "insumo").strip()
            unidade = (request.form.get("unidade") or "UN").strip()
            estoque_minimo = request.form.get("estoque_minimo")
            estoque_minimo = float(estoque_minimo) if estoque_minimo and str(estoque_minimo).strip() else None

            if not sku or not nome:
                flash("Informe SKU e nome.", "warning")
                return redirect(url_for("admin_produtos"))

            if Produto.query.filter(func.lower(Produto.sku) == sku.lower()).first():
                flash("SKU já existe.", "warning")
                return redirect(url_for("admin_produtos"))

            db.session.add(Produto(sku=sku, nome=nome, tipo=tipo, unidade=unidade, estoque_minimo=estoque_minimo))
            db.session.commit()
            flash("Produto criado.", "success")
            return redirect(url_for("admin_produtos"))

        produtos = Produto.query.order_by(Produto.nome).all()
        unidades = DEFAULT_UNIDADES
        return render_template("admin_produtos.html", produtos=produtos, unidades=unidades)

    @app.route("/admin/produtos/<int:produto_id>/edit", methods=["GET", "POST"])
    @admin_required
    def admin_produto_edit(produto_id):
        produto = Produto.query.get_or_404(produto_id)
        unidades = DEFAULT_UNIDADES
        if request.method == "POST":
            produto.sku = (request.form.get("sku") or "").strip()
            produto.nome = (request.form.get("nome") or "").strip()
            produto.tipo = (request.form.get("tipo") or "insumo").strip()
            produto.unidade = (request.form.get("unidade") or "UN").strip()
            em = request.form.get("estoque_minimo")
            produto.estoque_minimo = float(em) if em and str(em).strip() else None
            produto.ativo = request.form.get("ativo") == "1"
            produto.observacao = (request.form.get("observacao") or "").strip() or None
            db.session.commit()
            flash("Produto atualizado.", "success")
            return redirect(url_for("admin_produtos"))
        return render_template("admin_produto_edit.html", produto=produto, unidades=unidades)

    @app.route("/admin/produtos/<int:produto_id>/delete", methods=["POST"])
    @admin_required
    def admin_produto_delete(produto_id):
        produto = Produto.query.get_or_404(produto_id)
        db.session.delete(produto)
        db.session.commit()
        flash("Produto removido.", "info")
        return redirect(url_for("admin_produtos"))

    # -------- Patrimônio --------
    @app.route("/admin/patrimonio", methods=["GET", "POST"])
    @admin_required
    def admin_patrimonio():
        if request.method == "POST":
            categoria = (request.form.get("categoria") or "").strip()
            descricao = (request.form.get("descricao") or "").strip()
            status = (request.form.get("status") or "em_uso").strip()
            local_atual_id = request.form.get("local_atual_id", type=int)
            local_atual_id = local_atual_id if local_atual_id else None

            if not categoria or not descricao:
                flash("Informe categoria e descrição.", "warning")
                return redirect(url_for("admin_patrimonio"))

            codigo = next_patrimonio_codigo()
            it = PatrimonioItem(
                codigo=codigo,
                categoria=categoria,
                descricao=descricao,
                status=status,
                local_atual_id=local_atual_id
            )
            db.session.add(it)
            db.session.commit()

            mv = PatrimonioMov(
                item_id=it.id,
                tipo="entrada",
                origem_local_id=None,
                destino_local_id=local_atual_id,
                responsavel=None,
                observacao="Criado no sistema",
                projeto_id=None,
                qtd=1,
            )
            db.session.add(mv)
            db.session.commit()

            flash("Item patrimonial criado.", "success")
            return redirect(url_for("admin_patrimonio"))

        q = (request.args.get("q") or "").strip()
        status = (request.args.get("status") or "").strip()
        local_id = request.args.get("local_id", type=int)

        query = PatrimonioItem.query
        if q:
            like = f"%{q}%"
            query = query.filter(
                (PatrimonioItem.descricao.ilike(like)) |
                (PatrimonioItem.categoria.ilike(like)) |
                (PatrimonioItem.serial.ilike(like))
            )
        if status:
            query = query.filter(PatrimonioItem.status == status)
        if local_id:
            query = query.filter(PatrimonioItem.local_atual_id == local_id)

        itens = query.order_by(PatrimonioItem.id.desc()).limit(500).all()
        return render_template("admin_patrimonio.html", itens=itens, q=q, status=status, local_id=local_id)

    @app.route("/admin/patrimonio/<int:item_id>/edit", methods=["GET", "POST"])
    @admin_required
    def admin_patrimonio_edit(item_id):
        item = PatrimonioItem.query.get_or_404(item_id)
        if request.method == "POST":
            item.categoria = (request.form.get("categoria") or "").strip()
            item.descricao = (request.form.get("descricao") or "").strip()
            item.marca = (request.form.get("marca") or "").strip() or None
            item.modelo = (request.form.get("modelo") or "").strip() or None
            item.serial = (request.form.get("serial") or "").strip() or None
            item.responsavel = (request.form.get("responsavel") or "").strip() or None
            item.status = (request.form.get("status") or "em_uso").strip()
            lid = request.form.get("local_atual_id", type=int)
            item.local_atual_id = lid if lid else None

            # PROJETO ATUAL (corrigido)
            pid = request.form.get("projeto_id", type=int)
            item.projeto_atual_id = pid if pid else None

            val = request.form.get("valor")
            item.valor = float(val) if val and str(val).strip() else None
            item.observacao = (request.form.get("observacao") or "").strip() or None
            db.session.commit()
            flash("Item atualizado.", "success")
            return redirect(url_for("admin_patrimonio_edit", item_id=item.id))

        movs = PatrimonioMov.query.filter_by(item_id=item.id).order_by(PatrimonioMov.data.desc()).all()
        return render_template("admin_patrimonio_edit.html", item=item, movs=movs)

    @app.route("/admin/patrimonio/<int:item_id>/move", methods=["POST"])
    @admin_required
    def admin_patrimonio_move(item_id):
        item = PatrimonioItem.query.get_or_404(item_id)
        tipo = (request.form.get("tipo") or "transferencia").strip()
        origem = request.form.get("origem_local_id", type=int)
        destino = request.form.get("destino_local_id", type=int)
        responsavel = (request.form.get("responsavel") or "").strip() or None
        obs = (request.form.get("observacao") or "").strip() or None

        mv = PatrimonioMov(
            item_id=item.id,
            tipo=tipo,
            origem_local_id=origem or None,
            destino_local_id=destino or None,
            responsavel=responsavel,
            observacao=obs,
            projeto_id=None,
            qtd=1,
        )
        db.session.add(mv)

        if tipo in {"transferencia", "entrada"}:
            item.local_atual_id = destino or item.local_atual_id
        if tipo == "baixa":
            item.status = "baixado"
            item.projeto_atual_id = None
        if tipo == "manutencao":
            item.status = "manutencao"

        db.session.commit()
        flash("Movimentação registrada.", "success")
        return redirect(url_for("admin_patrimonio_edit", item_id=item.id))

    @app.route("/admin/patrimonio/<int:item_id>/delete", methods=["POST"])
    @admin_required
    def admin_patrimonio_delete(item_id):
        item = PatrimonioItem.query.get_or_404(item_id)
        PatrimonioMov.query.filter_by(item_id=item.id).delete()
        db.session.delete(item)
        db.session.commit()
        flash("Item removido.", "info")
        return redirect(url_for("admin_patrimonio"))

    @app.route("/admin/patrimonio/movs")
    @admin_required
    def admin_patrimonio_mov():
        Local2 = aliased(Local)

        rows = (db.session.query(
                    PatrimonioMov,
                    PatrimonioItem.codigo.label("item_codigo"),
                    PatrimonioItem.descricao.label("item_desc"),
                    Local.nome.label("origem_nome"),
                    Local2.nome.label("destino_nome"),
                    Projeto.nome.label("projeto_nome"),
                )
                .join(PatrimonioItem, PatrimonioMov.item_id == PatrimonioItem.id)
                .outerjoin(Local, PatrimonioMov.origem_local_id == Local.id)
                .outerjoin(Local2, PatrimonioMov.destino_local_id == Local2.id)
                .outerjoin(Projeto, PatrimonioMov.projeto_id == Projeto.id)
                .order_by(PatrimonioMov.data.desc())
                .limit(500)
                .all())

        movs = []
        for mv, item_codigo, item_desc, origem_nome, destino_nome, projeto_nome in rows:
            movs.append({
                "data": mv.data,
                "tipo": mv.tipo,
                "qtd": int(getattr(mv, "qtd", 1) or 1),
                "item": f"{item_codigo} — {item_desc}",
                "origem": origem_nome,
                "destino": destino_nome,
                "projeto": projeto_nome,
                "responsavel": mv.responsavel,
            })

        return render_template("admin_patrimonio_mov.html", movs=movs)

    @app.route("/admin/patrimonio/movs/add", methods=["POST"])
    @admin_required
    def admin_patrimonio_mov_add():
        item_id = request.form.get("item_id", type=int)
        tipo = (request.form.get("tipo") or "transferencia").strip()
        origem = request.form.get("origem_local_id", type=int)
        destino = request.form.get("destino_local_id", type=int)
        projeto_id = request.form.get("projeto_id", type=int)
        qtd = request.form.get("qtd", type=int) or 1
        try:
            qtd = max(1, int(qtd))
        except Exception:
            qtd = 1

        responsavel = (request.form.get("responsavel") or "").strip() or None
        obs = (request.form.get("observacao") or "").strip() or None

        item = PatrimonioItem.query.get_or_404(item_id)

        mv = PatrimonioMov(
            item_id=item.id,
            tipo=tipo,
            origem_local_id=origem if origem else None,
            destino_local_id=destino if destino else None,
            responsavel=responsavel,
            observacao=obs,
            projeto_id=projeto_id if projeto_id else None,
            qtd=qtd,
        )
        db.session.add(mv)

        if tipo in {"transferencia", "entrada"} and destino:
            item.local_atual_id = destino

        if tipo == "alocado":
            item.status = "alocado"
            item.projeto_atual_id = projeto_id if projeto_id else None
            if destino:
                item.local_atual_id = destino

        if tipo == "baixa":
            item.status = "baixado"
            item.projeto_atual_id = None

        if tipo == "manutencao":
            item.status = "manutencao"

        db.session.commit()
        flash("Movimentação registrada.", "success")
        return redirect(url_for("admin_patrimonio_mov"))

    # -------- Clientes / Projetos --------
    @app.route("/admin/clientes", methods=["GET", "POST"])
    @admin_required
    def admin_clientes():
        if request.method == "POST":
            nome = (request.form.get("nome") or "").strip()
            if not nome:
                flash("Informe o nome.", "warning")
                return redirect(url_for("admin_clientes"))
            c = Cliente(
                nome=nome,
                cnpj=(request.form.get("cnpj") or "").strip() or None,
                contato=(request.form.get("contato") or "").strip() or None,
                telefone=(request.form.get("telefone") or "").strip() or None,
                email=(request.form.get("email") or "").strip() or None,
            )
            db.session.add(c)
            db.session.commit()
            flash("Cliente criado.", "success")
            return redirect(url_for("admin_clientes"))
        clientes = Cliente.query.order_by(Cliente.nome).all()
        return render_template("admin_clientes.html", clientes=clientes)

    @app.route("/admin/clientes/<int:cliente_id>/delete", methods=["POST"])
    @admin_required
    def admin_cliente_delete(cliente_id):
        c = Cliente.query.get_or_404(cliente_id)
        db.session.delete(c)
        db.session.commit()
        flash("Cliente removido.", "info")
        return redirect(url_for("admin_clientes"))

    @app.route("/admin/projetos", methods=["GET", "POST"])
    @admin_required
    def admin_projetos():
        if request.method == "POST":
            nome = (request.form.get("nome") or "").strip()
            if not nome:
                flash("Informe o nome.", "warning")
                return redirect(url_for("admin_projetos"))
            pid = request.form.get("cliente_id", type=int)
            lid = request.form.get("local_id", type=int)
            pr = Projeto(
                nome=nome,
                cliente_id=pid if pid else None,
                local_id=lid if lid else None,
                status=(request.form.get("status") or "ativo").strip(),
                observacao=(request.form.get("observacao") or "").strip() or None,
            )
            db.session.add(pr)
            db.session.commit()
            flash("Projeto criado.", "success")
            return redirect(url_for("admin_projetos"))
        projetos = Projeto.query.order_by(Projeto.id.desc()).all()
        return render_template("admin_projetos.html", projetos=projetos)

    # -------- Estoque Mov --------
    @app.route("/admin/estoque/mov", methods=["GET", "POST"])
    @admin_required
    def admin_estoque_mov():
        if request.method == "POST":
            tipo = (request.form.get("tipo") or "").strip()
            produto_id = request.form.get("produto_id", type=int)
            qtd = float((request.form.get("qtd") or "0").replace(",", "."))
            dep_origem = request.form.get("deposito_origem_id", type=int)
            dep_dest = request.form.get("deposito_destino_id", type=int)
            dep_origem = dep_origem if dep_origem else None
            dep_dest = dep_dest if dep_dest else None
            doc = (request.form.get("documento_ref") or "").strip() or None
            obs = (request.form.get("observacao") or "").strip() or None

            try:
                apply_estoque_mov(tipo, produto_id, qtd, dep_origem, dep_dest, None, doc, obs)
                flash("Movimentação registrada.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Erro: {e}", "danger")
            return redirect(url_for("admin_estoque_mov"))

        movs = EstoqueMov.query.order_by(EstoqueMov.data.desc()).limit(500).all()
        return render_template("admin_estoque_mov.html", movs=movs)

    @app.route("/admin/inventario", methods=["GET", "POST"])
    @admin_required
    def admin_inventario():
        if request.method == "POST":
            deposito_id = request.form.get("deposito_id", type=int)
            produto_id = request.form.get("produto_id", type=int)
            novo_saldo = float((request.form.get("novo_saldo") or "0").replace(",", "."))
            try:
                apply_estoque_mov("ajuste", produto_id, novo_saldo, None, deposito_id, None,
                                  "inventario", "Ajuste por contagem (inventário)")
                flash("Ajuste aplicado.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Erro: {e}", "danger")
            return redirect(url_for("admin_inventario"))

        ajustes = EstoqueMov.query.filter_by(tipo="ajuste").order_by(EstoqueMov.data.desc()).limit(50).all()
        return render_template("admin_inventario.html", ajustes=ajustes)

    @app.route("/admin/relatorios")
    @admin_required
    def admin_relatorios():
        # paginação
        saldos_page = request.args.get("saldos_page", type=int) or 1
        minimo_page = request.args.get("minimo_page", type=int) or 1
        per_page = request.args.get("per_page", type=int) or 25
        per_page = max(5, min(per_page, 200))

        # Totais (cards)
        acabado_total = (db.session.query(func.coalesce(func.sum(EstoqueSaldo.saldo), 0))
                        .join(Produto, EstoqueSaldo.produto_id == Produto.id)
                        .filter(Produto.tipo == "acabado")
                        .scalar())
        insumo_total = (db.session.query(func.coalesce(func.sum(EstoqueSaldo.saldo), 0))
                        .join(Produto, EstoqueSaldo.produto_id == Produto.id)
                        .filter(Produto.tipo == "insumo")
                        .scalar())
        patrimonio_total = PatrimonioItem.query.count()

        # Saldos por tipo e depósito (resumo) — paginado
        base_q = (db.session.query(
                    Produto.tipo.label("tipo"),
                    Deposito.nome.label("deposito"),
                    func.coalesce(func.sum(EstoqueSaldo.saldo), 0).label("saldo"))
                .join(Produto, EstoqueSaldo.produto_id == Produto.id)
                .join(Deposito, EstoqueSaldo.deposito_id == Deposito.id)
                .group_by(Produto.tipo, Deposito.nome)
                .order_by(Produto.tipo, Deposito.nome))

        total_saldos = base_q.count()
        saldos_total_pages = max(1, (total_saldos + per_page - 1) // per_page)
        saldos_page = max(1, min(saldos_page, saldos_total_pages))
        by_tipo_deposito = (base_q.offset((saldos_page - 1) * per_page)
                                .limit(per_page)
                                .all())

        # Abaixo do estoque mínimo — paginado
        abaixo_q = (db.session.query(
                        Produto.sku, Produto.nome, Produto.tipo, Produto.estoque_minimo,
                        func.coalesce(func.sum(EstoqueSaldo.saldo), 0).label("saldo_total"))
                    .join(Produto, EstoqueSaldo.produto_id == Produto.id)
                    .filter(Produto.estoque_minimo != None)
                    .group_by(Produto.id)
                    .having(func.coalesce(func.sum(EstoqueSaldo.saldo), 0) < Produto.estoque_minimo)
                    .order_by(Produto.tipo, Produto.nome))

        total_minimo = abaixo_q.count()
        minimo_total_pages = max(1, (total_minimo + per_page - 1) // per_page)
        minimo_page = max(1, min(minimo_page, minimo_total_pages))
        abaixo_rows = abaixo_q.offset((minimo_page - 1) * per_page).limit(per_page).all()

        abaixo_minimo_page = [{
            "sku": r.sku,
            "nome": r.nome,
            "tipo": r.tipo,
            "saldo_total": float(r.saldo_total or 0),
            "minimo": float(r.estoque_minimo or 0),
        } for r in abaixo_rows]

        rel = {
            "acabado_total": float(acabado_total or 0),
            "insumo_total": float(insumo_total or 0),
            "patrimonio_total": int(patrimonio_total or 0),

            "by_tipo_deposito": [{"tipo": r.tipo, "deposito": r.deposito, "saldo": float(r.saldo or 0)} for r in by_tipo_deposito],
            "abaixo_minimo_page": abaixo_minimo_page,

            "per_page": per_page,
            "saldos_page": saldos_page,
            "saldos_total_pages": saldos_total_pages,
            "minimo_page": minimo_page,
            "minimo_total_pages": minimo_total_pages,
            "total_saldos": total_saldos,
            "total_minimo": total_minimo,

            # mantém compatibilidade com o trecho pg do template
            "pagination": {
                "page_saldos": saldos_page,
                "total_pages_saldos": saldos_total_pages,
                "page_min": minimo_page,
                "total_pages_min": minimo_total_pages,
            }
        }
        return render_template("admin_relatorios.html", rel=rel)
        
    # -------- Relatórios: detalhamento --------
    @app.route("/admin/relatorios/saldos")
    @admin_required
    def admin_relatorios_saldos():
        tipo = (request.args.get("tipo") or "todos").strip().lower()
        page = request.args.get("page", type=int) or 1
        per_page = request.args.get("per_page", type=int) or 50
        per_page = max(10, min(per_page, 200))

        q = (db.session.query(
                Produto.sku, Produto.nome, Produto.tipo, Produto.unidade,
                Deposito.nome.label("deposito"),
                func.coalesce(EstoqueSaldo.saldo, 0).label("saldo"))
             .join(Produto, EstoqueSaldo.produto_id == Produto.id)
             .join(Deposito, EstoqueSaldo.deposito_id == Deposito.id))

        if tipo in ("acabado", "insumo"):
            q = q.filter(Produto.tipo == tipo)

        total = q.count()
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = max(1, min(page, total_pages))

        rows = (q.order_by(Produto.nome, Deposito.nome)
                  .offset((page - 1) * per_page)
                  .limit(per_page)
                  .all())

        items = [{
            "sku": r.sku,
            "nome": r.nome,
            "tipo": r.tipo,
            "unidade": r.unidade,
            "deposito": r.deposito,
            "saldo": float(r.saldo or 0),
        } for r in rows]

        return render_template("admin_relatorios_saldos.html",
                               tipo=tipo, items=items,
                               page=page, per_page=per_page, total_pages=total_pages, total=total)

    @app.route("/admin/relatorios/patrimonio")
    @admin_required
    def admin_relatorios_patrimonio():
        page = request.args.get("page", type=int) or 1
        per_page = request.args.get("per_page", type=int) or 50
        per_page = max(10, min(per_page, 200))

        # Subquery: última movimentação por item (para obter qtd mais recente)
        sub_last = (db.session.query(
                        PatrimonioMov.item_id.label("item_id"),
                        func.max(PatrimonioMov.data).label("max_data")
                    )
                    .group_by(PatrimonioMov.item_id)
                    .subquery())

        last_mv = (db.session.query(
                        PatrimonioMov.item_id.label("item_id"),
                        PatrimonioMov.qtd.label("qtd"))
                .join(sub_last, (PatrimonioMov.item_id == sub_last.c.item_id) &
                                (PatrimonioMov.data == sub_last.c.max_data))
                .subquery())

        q = (db.session.query(
                PatrimonioItem.codigo.label("codigo"),
                PatrimonioItem.categoria.label("categoria"),
                PatrimonioItem.descricao.label("descricao"),
                PatrimonioItem.status.label("status"),
                Local.nome.label("local"),
                Municipio.nome.label("municipio"),
                Estado.sigla.label("uf"),
                func.coalesce(last_mv.c.qtd, 1).label("qtd"),
                Projeto.nome.label("projeto"),
            )
            .outerjoin(Local, PatrimonioItem.local_atual_id == Local.id)
            .outerjoin(Municipio, Local.municipio_id == Municipio.id)
            .outerjoin(Estado, Municipio.estado_id == Estado.id)
            .outerjoin(last_mv, last_mv.c.item_id == PatrimonioItem.id)
            .outerjoin(Projeto, PatrimonioItem.projeto_atual_id == Projeto.id)
        )

        total = q.count()
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = max(1, min(page, total_pages))

        rows = (q.order_by(PatrimonioItem.id.desc())
                .offset((page - 1) * per_page)
                .limit(per_page)
                .all())

        items = [{
            "codigo": r.codigo,
            "categoria": r.categoria,
            "descricao": r.descricao,
            "status": r.status,
            "local": r.local,
            "municipio": r.municipio,
            "uf": r.uf,
            "qtd": int(r.qtd or 1),
            "projeto": r.projeto,
        } for r in rows]

        return render_template(
            "admin_relatorios_patrimonio.html",
            items=items,
            page=page,
            per_page=per_page,
            total_pages=total_pages,
            total=total,
        )

    @app.route("/admin/relatorios/estoque-minimo")
    @admin_required
    def admin_relatorios_minimo():
        page = request.args.get("page", type=int) or 1
        per_page = request.args.get("per_page", type=int) or 50
        per_page = max(10, min(per_page, 200))

        q = (db.session.query(
                Produto.sku, Produto.nome, Produto.tipo, Produto.estoque_minimo,
                func.coalesce(func.sum(EstoqueSaldo.saldo), 0).label("saldo_total"))
             .join(Produto, EstoqueSaldo.produto_id == Produto.id)
             .filter(Produto.estoque_minimo != None)
             .group_by(Produto.id)
             .having(func.coalesce(func.sum(EstoqueSaldo.saldo), 0) < Produto.estoque_minimo)
             .order_by(Produto.tipo, Produto.nome))

        total = q.count()
        total_pages = max(1, (total + per_page - 1) // per_page)
        page = max(1, min(page, total_pages))

        rows = q.offset((page - 1) * per_page).limit(per_page).all()
        items = [{
            "sku": r.sku,
            "nome": r.nome,
            "tipo": r.tipo,
            "saldo_total": float(r.saldo_total or 0),
            "minimo": float(r.estoque_minimo or 0),
        } for r in rows]

        return render_template("admin_relatorios_minimo.html",
                               items=items, page=page, per_page=per_page, total_pages=total_pages, total=total)

    # -------- Exportações em planilha (XLSX) --------
    def _xlsx_response(filename: str, wb):
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return Response(
            bio.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @app.route("/admin/export/saldos.xlsx")
    @admin_required
    def export_saldos_xlsx():
        tipo = (request.args.get("tipo") or "todos").strip().lower()

        q = (db.session.query(
                Produto.sku, Produto.nome, Produto.tipo, Produto.unidade,
                Deposito.nome.label("deposito"),
                func.coalesce(EstoqueSaldo.saldo, 0).label("saldo"))
             .join(Produto, EstoqueSaldo.produto_id == Produto.id)
             .join(Deposito, EstoqueSaldo.deposito_id == Deposito.id))
        if tipo in ("acabado", "insumo"):
            q = q.filter(Produto.tipo == tipo)

        rows = q.order_by(Produto.nome, Deposito.nome).all()

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Saldos"
        ws.append(["SKU", "Produto", "Tipo", "Unidade", "Depósito", "Saldo"])
        for r in rows:
            ws.append([r.sku, r.nome, r.tipo, r.unidade, r.deposito, float(r.saldo or 0)])

        fn = f"saldos_{tipo}.xlsx" if tipo in ("acabado", "insumo") else "saldos_todos.xlsx"
        return _xlsx_response(fn, wb)

    @app.route("/admin/export/abaixo_minimo.xlsx")
    @admin_required
    def export_abaixo_minimo_xlsx():
        q = (db.session.query(
                Produto.sku, Produto.nome, Produto.tipo, Produto.estoque_minimo,
                func.coalesce(func.sum(EstoqueSaldo.saldo), 0).label("saldo_total"))
             .join(Produto, EstoqueSaldo.produto_id == Produto.id)
             .filter(Produto.estoque_minimo != None)
             .group_by(Produto.id)
             .having(func.coalesce(func.sum(EstoqueSaldo.saldo), 0) < Produto.estoque_minimo)
             .order_by(Produto.tipo, Produto.nome))
        rows = q.all()

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AbaixoMinimo"
        ws.append(["SKU", "Produto", "Tipo", "Saldo total", "Estoque mínimo"])
        for r in rows:
            ws.append([r.sku, r.nome, r.tipo, float(r.saldo_total or 0), float(r.estoque_minimo or 0)])

        return _xlsx_response("abaixo_estoque_minimo.xlsx", wb)

    @app.route("/admin/export/patrimonio.xlsx")
    @admin_required
    def export_patrimonio_xlsx():
        # Inclui projeto e qtd (última mov) no export
        sub_last = (db.session.query(
                        PatrimonioMov.item_id.label("item_id"),
                        func.max(PatrimonioMov.data).label("max_data")
                    )
                    .group_by(PatrimonioMov.item_id)
                    .subquery())

        last_mv = (db.session.query(
                        PatrimonioMov.item_id.label("item_id"),
                        PatrimonioMov.qtd.label("qtd"))
                   .join(sub_last, (PatrimonioMov.item_id == sub_last.c.item_id) & (PatrimonioMov.data == sub_last.c.max_data))
                   .subquery())

        rows = (db.session.query(
                    PatrimonioItem.codigo, PatrimonioItem.categoria, PatrimonioItem.descricao,
                    PatrimonioItem.status,
                    Local.nome.label("local"),
                    Municipio.nome.label("municipio"),
                    Estado.sigla.label("uf"),
                    func.coalesce(last_mv.c.qtd, 1).label("qtd"),
                    Projeto.nome.label("projeto"),
                )
                .outerjoin(Local, PatrimonioItem.local_atual_id == Local.id)
                .outerjoin(Municipio, Local.municipio_id == Municipio.id)
                .outerjoin(Estado, Municipio.estado_id == Estado.id)
                .outerjoin(last_mv, last_mv.c.item_id == PatrimonioItem.id)
                .outerjoin(Projeto, PatrimonioItem.projeto_atual_id == Projeto.id)
                .order_by(PatrimonioItem.codigo.desc())
                .all())

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Patrimonio"
        ws.append(["Código", "Categoria", "Descrição", "Status", "Quantidade", "Projeto", "Local", "Município", "UF"])
        for r in rows:
            ws.append([
                r.codigo, r.categoria, r.descricao, r.status,
                int(r.qtd or 1), r.projeto or "",
                r.local or "", r.municipio or "", r.uf or ""
            ])

        return _xlsx_response("patrimonio.xlsx", wb)

    # -------- Importações em massa (Produtos e Patrimônio) --------
    def _read_rows_csv(file_storage):
        stream = io.StringIO(file_storage.stream.read().decode("utf-8-sig"))
        reader = csv.DictReader(stream)
        return [{(k or "").strip().lower(): (v or "").strip() for k, v in row.items()} for row in reader]

    def _read_rows_xlsx(file_storage):
        import openpyxl
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip().lower() for h in rows[0]]
        out = []
        for r in rows[1:]:
            d = {}
            for i, h in enumerate(headers):
                d[h] = "" if i >= len(r) or r[i] is None else str(r[i]).strip()
            out.append(d)
        return out

    @app.route("/admin/produtos/import", methods=["POST"])
    @admin_required
    def admin_produtos_import():
        f = request.files.get("arquivo")
        if not f or not f.filename:
            flash("Envie um arquivo CSV ou XLSX.", "warning")
            return redirect(url_for("admin_produtos"))
        ext = f.filename.rsplit(".", 1)[1].lower()
        if ext not in ("csv", "xlsx"):
            flash("Arquivo deve ser CSV ou XLSX.", "danger")
            return redirect(url_for("admin_produtos"))
        try:
            rows = _read_rows_csv(f) if ext == "csv" else _read_rows_xlsx(f)
        except Exception as e:
            flash(f"Erro ao ler arquivo: {e}", "danger")
            return redirect(url_for("admin_produtos"))

        criados = 0
        atualizados = 0
        erros = 0
        for row in rows:
            try:
                sku = (row.get("sku") or "").strip()
                nome = (row.get("nome") or "").strip()
                tipo = (row.get("tipo") or "insumo").strip().lower() or "insumo"
                unidade = (row.get("unidade") or "UN").strip().upper() or "UN"
                em = (row.get("estoque_minimo") or "").strip()
                estoque_minimo = float(em.replace(",", ".")) if em else None
                ativo_raw = (row.get("ativo") or "").strip().lower()
                ativo = True if ativo_raw == "" else (ativo_raw in ("1", "true", "sim", "s", "yes", "y"))

                if not sku or not nome:
                    erros += 1
                    continue

                p = Produto.query.filter(func.lower(Produto.sku) == sku.lower()).first()
                if not p:
                    p = Produto(sku=sku, nome=nome, tipo=tipo, unidade=unidade, estoque_minimo=estoque_minimo, ativo=ativo)
                    db.session.add(p)
                    criados += 1
                else:
                    p.nome = nome
                    p.tipo = tipo
                    p.unidade = unidade
                    p.estoque_minimo = estoque_minimo
                    p.ativo = ativo
                    atualizados += 1
            except Exception:
                erros += 1

        db.session.commit()
        flash(f"Importação concluída. Criados: {criados}, atualizados: {atualizados}, erros: {erros}.",
              "success" if (criados or atualizados) else "warning")
        return redirect(url_for("admin_produtos"))

    @app.route("/admin/produtos/modelo.csv")
    @admin_required
    def admin_produtos_modelo():
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["sku", "nome", "tipo", "unidade", "estoque_minimo", "ativo"])
        w.writerow(["INS-001", "Papel A4 75g", "insumo", "CX", "5", "1"])
        w.writerow(["ACB-001", "Livro Matemática 6º", "acabado", "UN", "0", "1"])
        return Response(out.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": "attachment; filename=modelo_produtos.csv"})

    @app.route("/admin/patrimonio/import", methods=["POST"])
    @admin_required
    def admin_patrimonio_import():
        f = request.files.get("arquivo")
        if not f or not f.filename:
            flash("Envie um arquivo CSV ou XLSX.", "warning")
            return redirect(url_for("admin_patrimonio"))
        ext = f.filename.rsplit(".", 1)[1].lower()
        if ext not in ("csv", "xlsx"):
            flash("Arquivo deve ser CSV ou XLSX.", "danger")
            return redirect(url_for("admin_patrimonio"))
        try:
            rows = _read_rows_csv(f) if ext == "csv" else _read_rows_xlsx(f)
        except Exception as e:
            flash(f"Erro ao ler arquivo: {e}", "danger")
            return redirect(url_for("admin_patrimonio"))

        criados = 0
        erros = 0
        for row in rows:
            try:
                categoria = (row.get("categoria") or "").strip()
                descricao = (row.get("descricao") or "").strip()
                status = (row.get("status") or "em_uso").strip() or "em_uso"
                local_id = (row.get("local_id") or "").strip()
                local_id = int(local_id) if local_id else None
                marca = (row.get("marca") or "").strip() or None
                modelo = (row.get("modelo") or "").strip() or None
                serial = (row.get("serial") or "").strip() or None
                responsavel = (row.get("responsavel") or "").strip() or None
                val = (row.get("valor") or "").strip()
                valor = float(val.replace(",", ".")) if val else None

                if not categoria or not descricao:
                    erros += 1
                    continue

                codigo = next_patrimonio_codigo()
                it = PatrimonioItem(
                    codigo=codigo,
                    categoria=categoria,
                    descricao=descricao,
                    status=status,
                    local_atual_id=local_id,
                    marca=marca,
                    modelo=modelo,
                    serial=serial,
                    responsavel=responsavel,
                    valor=valor
                )
                db.session.add(it)
                db.session.flush()
                db.session.add(PatrimonioMov(
                    item_id=it.id,
                    tipo="entrada",
                    origem_local_id=None,
                    destino_local_id=local_id,
                    responsavel=responsavel,
                    observacao="Importação em massa",
                    projeto_id=None,
                    qtd=1,
                ))
                criados += 1
            except Exception:
                erros += 1

        db.session.commit()
        flash(f"Importação concluída. Criados: {criados}, erros: {erros}.",
              "success" if criados else "warning")
        return redirect(url_for("admin_patrimonio"))

    @app.route("/admin/patrimonio/modelo.csv")
    @admin_required
    def admin_patrimonio_modelo():
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(["categoria", "descricao", "status", "local_id", "marca", "modelo", "serial", "responsavel", "valor"])
        w.writerow(["TI", "Notebook Dell i5", "em_uso", "", "Dell", "Inspiron", "SN123", "João", "4500"])
        w.writerow(["Mobiliário", "Cadeira giratória", "em_uso", "", "", "", "", "", ""])
        return Response(out.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition": "attachment; filename=modelo_patrimonio.csv"})

    # ---------------- Init DB + migrations ----------------
    with app.app_context():
        db.create_all()
        _ensure_sqlite_columns(app)

    return app


# Gunicorn entrypoint
app = create_app()
application = app

if __name__ == "__main__":
    app.run(debug=True)